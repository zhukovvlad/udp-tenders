from datetime import date
from decimal import Decimal
from io import BytesIO
from itertools import groupby
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from crud.calculations import compute_export_rows
from crud.supplier_exclusions import get_excluded_supplier_ids
from database import get_db
from models import Project

router = APIRouter()

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = "000000", size: int = 10) -> Font:
    return Font(bold=bold, color=color, name="Calibri", size=size)


def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _safe_str(v: object) -> object:
    """Prevent Excel formula injection.

    Excel may interpret cells starting with =, +, - or @ as formulas. For any string
    whose first non-whitespace character is one of these, prefix an apostrophe.
    """
    if isinstance(v, str):
        s = v.lstrip()
        if s.startswith(("=", "+", "-", "@")):
            return "'" + v
    return v


_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Color palette
_C_HEADER_BG = "1F4E79"  # dark navy — project header block
_C_CLASS_BG = "2E75B6"  # medium blue — material-class section title
_C_COL_BG = "4472C4"  # lighter blue — column header row
_C_MONTH_BG = "DEEAF1"  # very light blue — month sub-header
_C_MONTH_TOTAL_BG = "BDD7EE"  # light blue — month subtotal row
_C_CLASS_TOTAL_BG = "9DC3E6"  # medium light blue — class grand total row
_C_TOTAL_BG = "D6E4F0"  # kept for reference (not used in sections)
_C_ODD = "FFFFFF"  # white — odd data rows
_C_EVEN = "EBF3FB"  # pale blue — even data rows
_C_RED_TEXT = "C00000"  # red — overpayment
_C_GREEN_TEXT = "375623"  # green — savings

_MONTH_NAMES_RU = [
    "", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]

_FMT_MONEY = '#,##0.00 "₽"'
_FMT_DATE = "DD.MM.YYYY"
_FMT_PCT = "+0.0%;-0.0%;0.0%"
_FMT_PCT_RATE = "0%"  # for vat_rate stored as decimal (0.20 → 20%)
_FMT_QTY = "#,##0.000"

# Column definitions: (header label, width, number_format, alignment)
# Col 1=A Дата, 2=B Номер, 3=C Поставщик
# Raw block:  4=D Кол-во по документу, 5=E Ед. изм. по документу
# Calc block: 6=F Расчётное кол-во,    7=G Базовая ед. изм.
# 8=H Базовая цена, 9=I Ставка НДС
# 10=J Материал без НДС, 11=K Доставка без НДС, 12=L Прочее без НДС, 13=M Итого без НДС (=J+K+L)
# 14=N Материал с НДС (=J*(1+I)), 15=O Доставка с НДС (=K*(1+I)), 16=P Прочее с НДС (=L*(1+I))
# 17=Q Итого с НДС (=N+O+P), 18=R Откл.% , 19=S Откл.₽
# 20=T Коридор %, 21=U Компенсация ₽
_COLUMNS = [
    ("Дата УПД",                  13, _FMT_DATE,      "center"),  # A  1
    ("Номер УПД",                 14, "@",            "left"),    # B  2
    ("Поставщик",                 30, "@",            "left"),    # C  3
    ("Кол-во по документу",       14, _FMT_QTY,       "right"),   # D  4  raw
    ("Ед. изм. по документу",     14, "@",            "center"),  # E  5  raw
    ("Расчётное кол-во",          14, _FMT_QTY,       "right"),   # F  6  normalized
    ("Базовая ед. изм.",          12, "@",            "center"),  # G  7  normalized
    ("Базовая цена",              16, _FMT_MONEY,     "right"),   # H  8
    ("Ставка НДС, %",             10, _FMT_PCT_RATE,  "center"),  # I  9
    ("Материал без НДС",          16, _FMT_MONEY,     "right"),   # J 10  static
    ("Доставка без НДС",          16, _FMT_MONEY,     "right"),   # K 11  static
    ("Прочее без НДС",            16, _FMT_MONEY,     "right"),   # L 12  static
    ("Итого без НДС",             16, _FMT_MONEY,     "right"),   # M 13  =J+K+L
    ("Материал с НДС",            16, _FMT_MONEY,     "right"),   # N 14  =J*(1+I)
    ("Доставка с НДС",            16, _FMT_MONEY,     "right"),   # O 15  =K*(1+I)
    ("Прочее с НДС",              16, _FMT_MONEY,     "right"),   # P 16  =L*(1+I)
    ("Итого с НДС",               16, _FMT_MONEY,     "right"),   # Q 17  =N+O+P
    ("Откл. от плана, %",         16, _FMT_PCT,       "right"),   # R 18  formula
    ("Откл. от плана, ₽",         16, _FMT_MONEY,     "right"),   # S 19  formula
    ("Коридор, %",                11, _FMT_PCT_RATE,  "center"),  # T 20  static
    ("Компенсация, ₽",            16, _FMT_MONEY,     "right"),   # U 21  Python value
]
_N_COLS = len(_COLUMNS)


def _dev_font(value, bold: bool = False, size: int = 10) -> Font:
    """Font with red/green colour based on deviation sign."""
    if value is None or value == 0:
        return _font(bold=bold, size=size)
    color = _C_RED_TEXT if value > 0 else _C_GREEN_TEXT
    return _font(bold=bold, color=color, size=size)


def _write_grand_total_row(
    ws,
    row_num: int,
    label: str,
    fill: PatternFill,
    label_font: Font,
    data_font: Font,
    data_ranges: list[tuple[int, int]],
    dev_total_py: float,
    comp_total: float | None = None,
) -> None:
    """Write the class-level grand total row across multiple non-contiguous month data ranges."""
    r = row_num

    def _c(col_idx, value, font=None, fmt=None, h="right"):
        cell = ws.cell(row=r, column=col_idx, value=value)
        cell.fill = fill
        cell.font = font or data_font
        cell.border = _BORDER
        cell.alignment = _align(h=h)
        if fmt:
            cell.number_format = fmt
        return cell

    _c(1, label, font=label_font, h="left")
    for ci in (2, 3, 4, 5, 7):   # B,C, raw-qty D, raw-unit E, base-unit G — blank/not aggregated
        _c(ci, None)

    sum_f = ",".join(f"F{s}:F{e}" for s, e in data_ranges)
    _c(6, f"=SUM({sum_f})", fmt=_FMT_QTY)        # F Расчётное кол-во
    _c(8, None)   # H Базовая цена — not averaged
    _c(9, None)   # I Ставка НДС — not averaged

    # Weighted averages: (Σ SUMPRODUCT(col, F)) / SUM(all F)
    for ci, cl in ((10, "J"), (11, "K"), (12, "L"), (14, "N"), (15, "O"), (16, "P")):
        sp = "+".join(f"SUMPRODUCT({cl}{s}:{cl}{e},F{s}:F{e})" for s, e in data_ranges)
        _c(ci, f'=IFERROR(({sp})/SUM({sum_f}),"")', fmt=_FMT_MONEY)

    _c(13, f"=J{r}+K{r}+L{r}", fmt=_FMT_MONEY)   # M Итого без НДС
    _c(17, f"=N{r}+O{r}+P{r}", fmt=_FMT_MONEY)   # Q Итого с НДС

    sum_s = ",".join(f"S{s}:S{e}" for s, e in data_ranges)
    _c(19, f'=IF(COUNT({sum_s})=0,"",SUM({sum_s}))', font=_dev_font(dev_total_py, bold=True, size=12), fmt=_FMT_MONEY)

    denom = "+".join(f"SUMPRODUCT((H{s}:H{e}>0)*H{s}:H{e}*F{s}:F{e})" for s, e in data_ranges)
    _c(18, f'=IFERROR(S{r}/({denom}),"")', font=_dev_font(dev_total_py, bold=True, size=12), fmt=_FMT_PCT)

    _c(20, None)  # T Коридор % — not aggregated at class level
    _c(21, comp_total, font=_dev_font(comp_total or 0, bold=True, size=12), fmt=_FMT_MONEY)

    ws.row_dimensions[r].height = 24


def _write_class_section(
    ws,
    class_name: str,
    rows: list[dict],
    start_row: int,
    comp_by_class_month: dict[tuple[int, int, int], dict] | None = None,
    material_class_id: int | None = None,
) -> int:
    """Write one material-class section with per-month breakdown.

    Layout per class:
      [Class header]
      [Column headers]
      [Month header — Январь 2025]  ← single row with aggregated SUMPRODUCT formulas
        row … row (data)
      [Month header — Февраль 2025]
        …
      [ИТОГО по <class>]
      (spacer)

    Columns 1–12 (A–L): static DB values.
    Columns 13–19 (M–S): Excel formulas (totals and deviations).
    Subtotal and grand-total rows also use SUMPRODUCT/SUM formulas.
    """
    cur = start_row

    # ── Class section header ─────────────────────────────────────────────────
    ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=_N_COLS)
    h = ws.cell(row=cur, column=1, value=_safe_str(class_name))
    h.fill = _fill(_C_CLASS_BG)
    h.font = _font(bold=True, color="FFFFFF", size=11)
    h.alignment = _align(h="left", v="center")
    ws.row_dimensions[cur].height = 20
    cur += 1

    # ── Column headers ───────────────────────────────────────────────────────
    for col_idx, (label, _, _, h_align) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=cur, column=col_idx, value=label)
        cell.fill = _fill(_C_COL_BG)
        cell.font = _font(bold=True, color="FFFFFF", size=9)
        cell.border = _BORDER
        cell.alignment = _align(h=h_align, wrap=True)
    ws.row_dimensions[cur].height = 30
    cur += 1

    # ── Month groups ─────────────────────────────────────────────────────────
    # rows are pre-sorted by date → groupby month is safe
    data_ranges: list[tuple[int, int]] = []  # (data_start, data_end) per month

    month_groups = [
        ((year, month), list(month_iter))
        for (year, month), month_iter in groupby(
            rows, key=lambda r: (r["invoice_date"].year, r["invoice_date"].month)
        )
    ]
    for group_idx, ((year, month), month_rows) in enumerate(month_groups):
        is_last_month = group_idx == len(month_groups) - 1
        month_label = f"{_MONTH_NAMES_RU[month]} {year}"

        # ── Month header row: label (A–C merged) + weighted-avg data (D–P) ─
        month_data_start = cur + 1
        month_data_end   = cur + len(month_rows)
        s, e = month_data_start, month_data_end
        rh = cur  # header row number

        # Python-side values for deviation font coloring (computed ahead of writing)
        month_dev = sum(
            row["deviation_amount"] for row in month_rows
            if row["deviation_amount"] is not None
        )

        hdr_fill = _fill(_C_MONTH_BG)
        hdr_font = _font(bold=True, color="1F4E79", size=11)

        ws.merge_cells(start_row=rh, start_column=1, end_row=rh, end_column=3)
        _mh = ws.cell(row=rh, column=1, value=month_label)
        _mh.fill = hdr_fill
        _mh.font = hdr_font
        _mh.border = _BORDER
        _mh.alignment = _align(h="left", v="center")
        for _ci in (2, 3):
            _mc = ws.cell(row=rh, column=_ci)
            _mc.fill = hdr_fill
            _mc.border = _BORDER

        def _hc(col_idx, value, font=None, fmt=None, h="right",
                _row=rh, _fill=hdr_fill, _font=hdr_font):
            cell = ws.cell(row=_row, column=col_idx, value=value)
            cell.fill = _fill
            cell.font = font or _font
            cell.border = _BORDER
            cell.alignment = _align(h=h)
            if fmt:
                cell.number_format = fmt

        _hc(6,  f"=SUM(F{s}:F{e})", fmt=_FMT_QTY)
        _hc(8,  None)   # H Базовая — not averaged
        _hc(9,  None)   # I Ставка НДС
        _hc(10, f'=IFERROR(SUMPRODUCT(J{s}:J{e},F{s}:F{e})/SUM(F{s}:F{e}),"")', fmt=_FMT_MONEY)
        _hc(11, f'=IFERROR(SUMPRODUCT(K{s}:K{e},F{s}:F{e})/SUM(F{s}:F{e}),"")', fmt=_FMT_MONEY)
        _hc(12, f'=IFERROR(SUMPRODUCT(L{s}:L{e},F{s}:F{e})/SUM(F{s}:F{e}),"")', fmt=_FMT_MONEY)
        _hc(13, f"=J{rh}+K{rh}+L{rh}", fmt=_FMT_MONEY)   # M Итого без НДС
        _hc(14, f'=IFERROR(SUMPRODUCT(N{s}:N{e},F{s}:F{e})/SUM(F{s}:F{e}),"")', fmt=_FMT_MONEY)
        _hc(15, f'=IFERROR(SUMPRODUCT(O{s}:O{e},F{s}:F{e})/SUM(F{s}:F{e}),"")', fmt=_FMT_MONEY)
        _hc(16, f'=IFERROR(SUMPRODUCT(P{s}:P{e},F{s}:F{e})/SUM(F{s}:F{e}),"")', fmt=_FMT_MONEY)
        _hc(17, f"=N{rh}+O{rh}+P{rh}", fmt=_FMT_MONEY)   # Q Итого с НДС
        _hc(19, f'=IF(COUNT(S{s}:S{e})=0,"",SUM(S{s}:S{e}))', font=_dev_font(month_dev, bold=True), fmt=_FMT_MONEY)
        _hc(18, f'=IFERROR(S{rh}/SUMPRODUCT((H{s}:H{e}>0)*H{s}:H{e}*F{s}:F{e}),"")',
            font=_dev_font(month_dev, bold=True), fmt=_FMT_PCT)
        # Columns T/U: corridor % and monthly compensation (Python values, nonlinear → not formulaic)
        month_key = (material_class_id, year, month)
        month_comp = (comp_by_class_month or {}).get(month_key, {})
        _corridor = month_comp.get("corridor_pct")
        _comp_amt = month_comp.get("compensation_amount")
        _hc(20, (_corridor / Decimal("100")) if _corridor is not None else None, fmt=_FMT_PCT_RATE)
        _hc(21, _comp_amt, font=_dev_font(_comp_amt or 0, bold=True), fmt=_FMT_MONEY)
        ws.row_dimensions[rh].height = 18
        cur += 1

        # Data rows
        for i, r in enumerate(month_rows):
            row_fill = _fill(_C_ODD if i % 2 == 0 else _C_EVEN)
            row_font = _font(color="000000")

            # Cols 1–12 (A–L): static DB values
            for col_idx, val in enumerate([
                r["invoice_date"],                # A 1
                _safe_str(r["invoice_number"]),   # B 2
                _safe_str(r["supplier_name"]),    # C 3
                r["raw_qty"],                     # D 4  Кол-во по документу
                _safe_str(r["raw_unit"]),         # E 5  Ед. изм. по документу
                r["qty"],                         # F 6  Расчётное кол-во
                _safe_str(r["unit_symbol"]),      # G 7  Базовая ед. изм.
                r["ref_price"],                   # H 8  Базовая цена
                r["vat_rate"],                    # I 9  Ставка НДС
                r["mat_per_m3_excl_vat"],         # J 10 Материал без НДС
                r["delivery_per_m3_excl_vat"],    # K 11 Доставка без НДС
                r["other_per_m3_excl_vat"],       # L 12 Прочее без НДС
            ], start=1):
                _, _, num_fmt, h_align = _COLUMNS[col_idx - 1]
                cell = ws.cell(row=cur, column=col_idx, value=val)
                cell.fill = row_fill
                cell.font = row_font
                cell.border = _BORDER
                cell.alignment = _align(h=h_align)
                if num_fmt:
                    cell.number_format = num_fmt

            n = cur

            def _dc(col_idx, value, font=None, fmt=_FMT_MONEY,
                    _fill=row_fill, _font=row_font, _row=n):
                cell = ws.cell(row=_row, column=col_idx, value=value)
                cell.fill = _fill
                cell.font = font or _font
                cell.border = _BORDER
                cell.alignment = _align(h="right")
                if fmt:
                    cell.number_format = fmt

            _dc(13, f"=J{n}+K{n}+L{n}")               # M Итого без НДС
            _dc(14, f"=J{n}*(1+I{n})")                # N Материал с НДС
            _dc(15, f"=K{n}*(1+I{n})")                # O Доставка с НДС
            _dc(16, f"=L{n}*(1+I{n})")                # P Прочее с НДС
            _dc(17, f"=N{n}+O{n}+P{n}")               # Q Итого с НДС
            _dc(18, f'=IFERROR(IF(H{n}>0,(Q{n}-H{n})/H{n},""),"")',
                font=_dev_font(r["deviation_pct"]), fmt=_FMT_PCT)   # R Откл.%
            _dc(19, f'=IFERROR(IF(H{n}>0,(Q{n}-H{n})*F{n},""),"")',
                font=_dev_font(r["deviation_amount"]))             # S Откл.₽

            ws.row_dimensions[cur].height = 16
            cur += 1

        data_ranges.append((month_data_start, month_data_end))

        # Thin separator row between months (not after the last one)
        if not is_last_month:
            for _ci in range(1, _N_COLS + 1):
                _sc = ws.cell(row=cur, column=_ci)
                _sc.fill = _fill(_C_MONTH_TOTAL_BG)
            ws.row_dimensions[cur].height = 6
            cur += 1

    # ── Class grand total ────────────────────────────────────────────────────
    grand_dev = sum(r["deviation_amount"] for r in rows if r["deviation_amount"] is not None)
    class_comp_total: float | None = None
    if comp_by_class_month and material_class_id is not None:
        monthly_comp_amounts = [
            v["compensation_amount"]
            for (cid, _y, _m), v in comp_by_class_month.items()
            if cid == material_class_id and v["compensation_amount"] is not None
        ]
        if monthly_comp_amounts:
            class_comp_total = sum(monthly_comp_amounts)
    _write_grand_total_row(
        ws, cur,
        label=f"ИТОГО по {class_name}",
        fill=_fill(_C_CLASS_TOTAL_BG),
        label_font=_font(bold=True, color="1F4E79", size=12),
        data_font=_font(bold=True, color="1F4E79", size=12),
        data_ranges=data_ranges,
        dev_total_py=grand_dev,
        comp_total=class_comp_total,
    )
    cur += 1

    # Spacer
    cur += 1
    return cur


# ---------------------------------------------------------------------------
# Endpoint
# ---------------------------------------------------------------------------

@router.get("/excel")
def export_excel(
    project_id: int,
    period_start: date | None = None,
    period_end: date | None = None,
    material_class_id: int | None = None,
    db: Session = Depends(get_db),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Проект не найден")

    excluded = get_excluded_supplier_ids(db, project_id)
    rows = compute_export_rows(
        db, project_id, period_start, period_end, material_class_id,
        excluded_supplier_ids=excluded or None,
    )

    from crud.calculations import compute_calculations  # noqa: PLC0415
    monthly_rows = compute_calculations(
        db, project_id, period_start, period_end, material_class_id,
        excluded_supplier_ids=excluded or None,
    )
    # (class_id, year, month) → {"corridor_pct": float|None, "compensation_amount": float|None}
    comp_by_class_month: dict[tuple[int, int, int], dict] = {
        (m["material_class_id"], m["period_start"].year, m["period_start"].month): {
            "corridor_pct": m["corridor_pct"],
            "compensation_amount": m["compensation_amount"],
        }
        for m in monthly_rows
    }

    # Actual displayed period from data
    if rows:
        display_start = period_start or min(r["invoice_date"] for r in rows)
        display_end = period_end or max(r["invoice_date"] for r in rows)
    else:
        display_start = period_start or date.today().replace(day=1)
        display_end = period_end or date.today()

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"
    ws.sheet_view.showGridLines = False

    # Set column widths
    for col_idx, (_, width, _, _) in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    cur = 1  # current row pointer

    # ── Project info block ────────────────────────────────────────────────────
    info_fill = _fill(_C_HEADER_BG)
    info_lines = [
        _safe_str(project.name),
        _safe_str(project.contract_number or ""),
        f"Период: {display_start.strftime('%d.%m.%Y')} — {display_end.strftime('%d.%m.%Y')}",
        f"Сформировано: {date.today().strftime('%d.%m.%Y')}",
    ]
    info_fonts = [
        _font(bold=True, color="FFFFFF", size=14),
        _font(color="BDD7EE", size=10),
        _font(color="FFFFFF", size=10),
        _font(color="9DC3E6", size=9),
    ]
    info_heights = [24, 16, 16, 14]

    for text, font, height in zip(info_lines, info_fonts, info_heights, strict=True):
        ws.merge_cells(
            start_row=cur, start_column=1,
            end_row=cur, end_column=_N_COLS,
        )
        cell = ws.cell(row=cur, column=1, value=text)
        cell.fill = info_fill
        cell.font = font
        cell.alignment = _align(h="left", v="center")
        ws.row_dimensions[cur].height = height
        cur += 1

    cur += 1  # blank spacer after header

    # ── Material class sections ───────────────────────────────────────────────
    if not rows:
        ws.merge_cells(
            start_row=cur, start_column=1, end_row=cur, end_column=_N_COLS
        )
        ws.cell(row=cur, column=1, value="Нет данных за выбранный период").font = _font(
            color="888888", size=10
        )
    else:
        for class_name, group in groupby(rows, key=lambda r: r["material_class_name"]):
            group_rows = list(group)
            cur = _write_class_section(
                ws, class_name, group_rows, cur,
                comp_by_class_month=comp_by_class_month,
                material_class_id=group_rows[0]["material_class_id"],
            )

    # ── Footer ────────────────────────────────────────────────────────────────
    footer_cell = ws.cell(
        row=cur + 1,
        column=1,
        value=(
            "* Стоимость доставки и прочих включений распределена пропорционально объёму м³ "
            "каждого класса материала в рамках каждой СФ.  "
            "** Компенсация считается от средней цены за месяц и показана в строках месяца и итога; "
            "по отдельным СФ не определяется."
        ),
    )
    footer_cell.font = _font(color="888888", size=8)
    footer_cell.alignment = _align(h="left", wrap=True)
    ws.merge_cells(
        start_row=cur + 1, start_column=1, end_row=cur + 1, end_column=_N_COLS
    )
    ws.row_dimensions[cur + 1].height = 24

    # ── Freeze panes ─────────────────────────────────────────────────────────
    # (no freeze — sections have their own headers)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = "".join("-" if ch in "\\/" + ':*?"<>|\r\n' else ch for ch in project.name).strip(" .-")
    filename = f"отчёт_{safe_name or project.id}_{display_start}_{display_end}.xlsx"
    encoded = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )
