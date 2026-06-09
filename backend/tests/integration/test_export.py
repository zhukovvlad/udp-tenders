"""Tests for the Excel export: compute_export_rows() logic and /api/export/excel endpoint.

Layout overview (21 columns A–U):
  A=Date  B=Number  C=Supplier
  D=RawQty  E=RawUnit  F=CalcQty (normalized)  G=BaseUnit
  H=RefPrice  I=VATrate
  J=MatExcl  K=DelivExcl  L=OtherExcl  M=formula:J+K+L
  N=formula:J*(1+I)  O=formula:K*(1+I)  P=formula:L*(1+I)  Q=formula:N+O+P
  R=formula:deviation%  S=formula:deviation₽
  T=corridor%  U=compensation₽
"""
from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from crud.calculations import compute_export_rows

# ---------------------------------------------------------------------------
# Sheet-scanning helpers
# ---------------------------------------------------------------------------

def _all_values(ws) -> list:
    return [
        ws.cell(row=r, column=c).value
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
    ]


def _find_data_row(ws) -> int | None:
    """Return 1-based row of the first row whose col D (4, raw_qty) holds a plain number > 0."""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=4).value
        if isinstance(v, int | float) and not isinstance(v, bool) and v > 0:
            return r
    return None


# ---------------------------------------------------------------------------
# Section 1: compute_export_rows() — business-logic unit tests (DB-backed)
# ---------------------------------------------------------------------------

class TestComputeExportRows:

    # ── Empty / boundary cases ───────────────────────────────────────────────

    def test_empty_when_no_invoices(self, db_session, factories):
        project = factories.ProjectFactory.create()
        assert compute_export_rows(db_session, project.id) == []

    def test_empty_when_period_excludes_all_invoices(self, db_session, factories):
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 1, 15))
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc, quantity=10.0)

        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        assert rows == []

    def test_empty_when_only_delivery_items_no_base(self, db_session, factories):
        """Invoice has only delivery items — no base material → no rows returned."""
        project = factories.ProjectFactory.create()
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10))
        # item_type="delivery" is picked up by delivery bucket, never base_rows
        factories.InvoiceItemFactory.create(
            invoice=inv, item_type="delivery",
            quantity=1.0, unit_price=50_000.0,
            amount=50_000.0, vat_amount=10_000.0,
        )
        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        assert rows == []

    # ── Row shape and basic arithmetic ──────────────────────────────────────

    def test_basic_row_has_all_expected_keys(self, db_session, factories):
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(name="В25", calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10), vat_rate=20.0)
        factories.InvoiceItemFactory.create(
            invoice=inv, material_class=mc,
            quantity=100.0, unit_price=5000.0,
            amount=500_000.0, vat_amount=100_000.0,
        )
        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        assert len(rows) == 1
        r = rows[0]
        required_keys = {
            "material_class_id", "material_class_name", "invoice_id",
            "invoice_date", "invoice_number", "supplier_name",
            "raw_qty", "raw_unit", "qty", "unit_symbol",
            "ref_price", "vat_rate",
            "mat_per_m3_excl_vat", "mat_per_m3",
            "delivery_per_m3_excl_vat", "delivery_per_m3",
            "other_per_m3_excl_vat", "other_per_m3",
            "total_per_m3", "deviation_pct", "deviation_amount",
        }
        assert required_keys <= r.keys()

    def test_mat_per_m3_excl_vat_equals_amount_over_qty(self, db_session, factories):
        """mat_per_m3_excl_vat = amount / qty (before VAT)."""
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10), vat_rate=20.0)
        factories.InvoiceItemFactory.create(
            invoice=inv, material_class=mc,
            quantity=100.0, unit_price=5000.0,
            amount=500_000.0, vat_amount=100_000.0,
        )
        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        r = rows[0]
        assert float(r["mat_per_m3_excl_vat"]) == pytest.approx(5000.0)
        # with VAT: (500_000 + 100_000) / 100 = 6000
        assert float(r["mat_per_m3"]) == pytest.approx(6000.0)
        assert float(r["vat_rate"]) == pytest.approx(0.20)

    def test_vat_amount_fallback_to_invoice_rate_when_null(self, db_session, factories):
        """When InvoiceItem.vat_amount is NULL, VAT is computed via invoice.vat_rate (20%)."""
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10), vat_rate=20.0)
        factories.InvoiceItemFactory.create(
            invoice=inv, material_class=mc,
            quantity=10.0, unit_price=6000.0,
            amount=60_000.0, vat_amount=None,
        )
        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        r = rows[0]
        assert float(r["mat_per_m3_excl_vat"]) == pytest.approx(6000.0)
        # computed VAT = 60_000 × 20% = 12_000 → total = 72_000 / 10 = 7200
        assert float(r["mat_per_m3"]) == pytest.approx(7200.0)

    def test_vat_rate_none_falls_back_to_20_percent(self, db_session, factories):
        """COALESCE(vat_rate, 20.0) — invoice with NULL vat_rate treated as 20%."""
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10), vat_rate=None)
        factories.InvoiceItemFactory.create(
            invoice=inv, material_class=mc,
            quantity=10.0, unit_price=6000.0,
            amount=60_000.0, vat_amount=None,
        )
        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        assert float(rows[0]["mat_per_m3"]) == pytest.approx(7200.0)

    def test_total_per_m3_equals_mat_plus_delivery_plus_other(self, db_session, factories):
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10), vat_rate=20.0)
        factories.InvoiceItemFactory.create(
            invoice=inv, material_class=mc, item_type="material",
            quantity=50.0, unit_price=5000.0, amount=250_000.0, vat_amount=50_000.0,
        )
        factories.InvoiceItemFactory.create(
            invoice=inv, item_type="delivery",
            quantity=1.0, unit_price=10_000.0, amount=10_000.0, vat_amount=2_000.0,
        )
        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        r = rows[0]
        assert float(r["total_per_m3"]) == pytest.approx(
            float(r["mat_per_m3"] + r["delivery_per_m3"] + r["other_per_m3"])
        )

    # ── Delivery proportional allocation ────────────────────────────────────

    def test_delivery_zero_when_no_delivery_items(self, db_session, factories):
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10))
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc, quantity=10.0)
        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        assert float(rows[0]["delivery_per_m3"]) == pytest.approx(0.0)
        assert float(rows[0]["delivery_per_m3_excl_vat"]) == pytest.approx(0.0)

    def test_delivery_allocated_proportionally_to_base_qty(self, db_session, factories):
        """Delivery is split 75/25 when base qtys are 75 and 25 m³."""
        project = factories.ProjectFactory.create()
        mc1 = factories.MaterialClassFactory.create(name="В25", calc_role="base")
        mc2 = factories.MaterialClassFactory.create(name="В30", calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10), vat_rate=20.0)
        factories.InvoiceItemFactory.create(
            invoice=inv, material_class=mc1, item_type="material",
            quantity=75.0, unit_price=5000.0, amount=375_000.0, vat_amount=75_000.0,
        )
        factories.InvoiceItemFactory.create(
            invoice=inv, material_class=mc2, item_type="material",
            quantity=25.0, unit_price=5000.0, amount=125_000.0, vat_amount=25_000.0,
        )
        # delivery total: 100_000 excl + 20_000 VAT = 120_000 with VAT
        factories.InvoiceItemFactory.create(
            invoice=inv, item_type="delivery",
            quantity=1.0, unit_price=100_000.0, amount=100_000.0, vat_amount=20_000.0,
        )
        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        row25 = next(r for r in rows if r["material_class_name"] == "В25")
        row30 = next(r for r in rows if r["material_class_name"] == "В30")

        # Each class: per-m³ delivery = same rate regardless of share (same unit price)
        # mc1: 75% × 120_000 / 75 = 1200; mc2: 25% × 120_000 / 25 = 1200
        assert float(row25["delivery_per_m3"]) == pytest.approx(1200.0)
        assert float(row30["delivery_per_m3"]) == pytest.approx(1200.0)
        # excl VAT: 75% × 100_000 / 75 = 1000
        assert float(row25["delivery_per_m3_excl_vat"]) == pytest.approx(1000.0)
        assert float(row30["delivery_per_m3_excl_vat"]) == pytest.approx(1000.0)

    # ── Additive (calc_role="additive") allocation ───────────────────────────

    def test_other_zero_when_no_additive_items(self, db_session, factories):
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10))
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc, quantity=10.0)
        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        assert float(rows[0]["other_per_m3"]) == pytest.approx(0.0)
        assert float(rows[0]["other_per_m3_excl_vat"]) == pytest.approx(0.0)

    def test_additive_allocated_to_base_class_proportionally(self, db_session, factories):
        """Additive (calc_role='additive', item_type='material') is allocated as 'other'."""
        project = factories.ProjectFactory.create()
        mc_base = factories.MaterialClassFactory.create(name="В25", calc_role="base")
        mc_add = factories.MaterialClassFactory.create(name="Пластификатор", calc_role="additive")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10), vat_rate=20.0)
        factories.InvoiceItemFactory.create(
            invoice=inv, material_class=mc_base, item_type="material",
            quantity=100.0, unit_price=5000.0, amount=500_000.0, vat_amount=100_000.0,
        )
        # additive: 20_000 excl + 4_000 VAT
        factories.InvoiceItemFactory.create(
            invoice=inv, material_class=mc_add, item_type="material",
            quantity=1.0, unit_price=20_000.0, amount=20_000.0, vat_amount=4_000.0,
        )
        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        # Only base class appears; additive is folded in
        assert len(rows) == 1
        r = rows[0]
        # mc_base is the only base class → share = 1.0 → all additive goes to it
        # other_per_m3 = (20_000 + 4_000) / 100 = 240
        assert float(r["other_per_m3"]) == pytest.approx(240.0)
        assert float(r["other_per_m3_excl_vat"]) == pytest.approx(200.0)  # 20_000 / 100

    # ── Deviation calculation ────────────────────────────────────────────────

    def test_deviation_none_when_no_ref_price(self, db_session, factories):
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10))
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc, quantity=10.0)
        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        r = rows[0]
        assert r["ref_price"] is None
        assert r["deviation_pct"] is None
        assert r["deviation_amount"] is None

    def test_deviation_none_when_ref_price_outside_invoice_date(self, db_session, factories):
        """Reference price period doesn't include invoice date → treated as no ref price."""
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        factories.ReferencePriceFactory.create(
            project=project, material_class=mc, price=7000.0,
            period_start=date(2026, 2, 1), period_end=date(2026, 2, 28),
        )
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10))
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc, quantity=10.0)
        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        assert rows[0]["ref_price"] is None
        assert rows[0]["deviation_pct"] is None

    def test_deviation_zero_when_total_equals_ref_price(self, db_session, factories):
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        factories.ReferencePriceFactory.create(
            project=project, material_class=mc, price=6000.0,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10), vat_rate=20.0)
        # total_per_m3 = (50_000 + 10_000) / 10 = 6000
        factories.InvoiceItemFactory.create(
            invoice=inv, material_class=mc,
            quantity=10.0, unit_price=5000.0, amount=50_000.0, vat_amount=10_000.0,
        )
        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        r = rows[0]
        assert r["ref_price"] == 6000.0
        assert float(r["deviation_pct"]) == pytest.approx(0.0)
        assert float(r["deviation_amount"]) == pytest.approx(0.0)

    def test_deviation_positive_and_correct_when_above_plan(self, db_session, factories):
        """total_per_m3 = 7200 vs ref = 5000 → dev = +44%, dev₽ = (7200-5000)×10 = 22_000."""
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        factories.ReferencePriceFactory.create(
            project=project, material_class=mc, price=5000.0,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10), vat_rate=20.0)
        factories.InvoiceItemFactory.create(
            invoice=inv, material_class=mc,
            quantity=10.0, unit_price=6000.0, amount=60_000.0, vat_amount=12_000.0,
        )
        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        r = rows[0]
        assert float(r["deviation_pct"]) == pytest.approx(44.0)
        assert float(r["deviation_amount"]) == pytest.approx(22_000.0)

    def test_deviation_negative_when_below_plan(self, db_session, factories):
        """total_per_m3 = 4800 vs ref = 6000 → dev = -20%, dev₽ = (4800-6000)×10 = -12_000."""
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        factories.ReferencePriceFactory.create(
            project=project, material_class=mc, price=6000.0,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10), vat_rate=20.0)
        # total = (40_000 + 8_000) / 10 = 4800
        factories.InvoiceItemFactory.create(
            invoice=inv, material_class=mc,
            quantity=10.0, unit_price=4000.0, amount=40_000.0, vat_amount=8_000.0,
        )
        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        r = rows[0]
        assert float(r["deviation_pct"]) == pytest.approx(-20.0)
        assert float(r["deviation_amount"]) == pytest.approx(-12_000.0)

    # ── Filtering and sorting ────────────────────────────────────────────────

    def test_material_class_id_filter_excludes_other_classes(self, db_session, factories):
        project = factories.ProjectFactory.create()
        mc1 = factories.MaterialClassFactory.create(name="В25", calc_role="base")
        mc2 = factories.MaterialClassFactory.create(name="В30", calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10))
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc1, quantity=10.0)
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc2, quantity=5.0)
        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
            material_class_id=mc1.id,
        )
        assert all(r["material_class_name"] == "В25" for r in rows)
        assert len(rows) == 1

    def test_sorted_by_class_name_then_invoice_date(self, db_session, factories):
        project = factories.ProjectFactory.create()
        mc_b30 = factories.MaterialClassFactory.create(name="В30", calc_role="base")
        mc_b25 = factories.MaterialClassFactory.create(name="В25", calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv_late = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 20), number="СФ-2")
        inv_early = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 5), number="СФ-1")
        for inv in (inv_late, inv_early):
            factories.InvoiceItemFactory.create(invoice=inv, material_class=mc_b25, quantity=10.0)
            factories.InvoiceItemFactory.create(invoice=inv, material_class=mc_b30, quantity=10.0)

        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        names = [r["material_class_name"] for r in rows]
        assert names == ["В25", "В25", "В30", "В30"]
        b25_rows = [r for r in rows if r["material_class_name"] == "В25"]
        assert b25_rows[0]["invoice_date"] < b25_rows[1]["invoice_date"]

    def test_period_auto_resolved_from_invoice_dates(self, db_session, factories):
        """Without explicit period, bounds are detected from min/max invoice date."""
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 4, 15))
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc, quantity=10.0)

        rows = compute_export_rows(db_session, project.id)
        assert len(rows) == 1
        assert rows[0]["invoice_date"] == date(2026, 4, 15)

    def test_multiple_invoices_same_class_produce_separate_rows(self, db_session, factories):
        """Two invoices for the same class → two separate rows, not merged."""
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv1 = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 5))
        inv2 = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 20))
        factories.InvoiceItemFactory.create(invoice=inv1, material_class=mc, quantity=50.0)
        factories.InvoiceItemFactory.create(invoice=inv2, material_class=mc, quantity=80.0)

        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        assert len(rows) == 2
        qtys = {r["qty"] for r in rows}
        assert qtys == {50.0, 80.0}

    # ── Supplier exclusion ───────────────────────────────────────────────────

    def test_excluded_supplier_rows_omitted(self, db_session, factories):
        """compute_export_rows не включает инвойсы исключённых поставщиков.

        Инвойсы без supplier_id (supplier_id IS NULL) при этом остаются —
        фильтр «OR supplier_id IS NULL» в excluded_supplier_ids.
        """
        from crud.supplier_exclusions import get_excluded_supplier_ids, set_supplier_excluded

        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        included_supplier = factories.SupplierFactory.create()
        excluded_supplier = factories.SupplierFactory.create()

        doc = factories.DocumentFactory.create(project=project)

        # Инвойс включённого поставщика
        inv_inc = factories.InvoiceFactory.create(
            document=doc, date=date(2026, 3, 5), supplier_id=included_supplier.id,
        )
        factories.InvoiceItemFactory.create(
            invoice=inv_inc, material_class=mc, item_type="material",
            quantity=10.0, unit_price=5000.0, amount=50000.0, vat_amount=10000.0,
        )

        # Инвойс исключённого поставщика
        inv_exc = factories.InvoiceFactory.create(
            document=doc, date=date(2026, 3, 10), supplier_id=excluded_supplier.id,
        )
        factories.InvoiceItemFactory.create(
            invoice=inv_exc, material_class=mc, item_type="material",
            quantity=20.0, unit_price=8000.0, amount=160000.0, vat_amount=32000.0,
        )

        # Инвойс без supplier_id — должен остаться в результате
        inv_null = factories.InvoiceFactory.create(
            document=doc, date=date(2026, 3, 15), supplier_id=None,
        )
        factories.InvoiceItemFactory.create(
            invoice=inv_null, material_class=mc, item_type="material",
            quantity=5.0, unit_price=6000.0, amount=30000.0, vat_amount=6000.0,
        )

        # Регистрируем исключение через CRUD — проверяем реальный путь получения excluded_supplier_ids
        set_supplier_excluded(db_session, project.id, excluded_supplier.id, excluded=True)
        excluded = get_excluded_supplier_ids(db_session, project.id)

        rows = compute_export_rows(
            db_session, project.id,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
            excluded_supplier_ids=excluded or None,
        )

        # Должны быть только 2 инвойса (включённый + без поставщика)
        assert len(rows) == 2
        invoice_ids = {r["invoice_id"] for r in rows}
        assert inv_inc.id in invoice_ids
        assert inv_null.id in invoice_ids
        assert inv_exc.id not in invoice_ids


# ---------------------------------------------------------------------------
# Section 2: Excel endpoint — HTTP + workbook structure
# ---------------------------------------------------------------------------

class TestExportEndpoint:

    # ── HTTP basics ──────────────────────────────────────────────────────────

    def test_returns_200_and_xlsx_content_type(self, client, factories):
        project = factories.ProjectFactory.create()
        resp = client.get(f"/api/export/excel?project_id={project.id}")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]

    def test_content_disposition_is_attachment_with_xlsx(self, client, factories):
        project = factories.ProjectFactory.create()
        resp = client.get(f"/api/export/excel?project_id={project.id}")
        cd = resp.headers.get("content-disposition", "")
        assert "attachment" in cd
        assert ".xlsx" in cd

    def test_unknown_project_returns_404(self, client):
        resp = client.get("/api/export/excel?project_id=999999")
        assert resp.status_code == 404
        assert resp.json()["detail"] == "Проект не найден"

    # ── No-data case ─────────────────────────────────────────────────────────

    def test_empty_period_renders_no_data_message(self, client, factories):
        project = factories.ProjectFactory.create()
        resp = client.get(
            f"/api/export/excel?project_id={project.id}"
            "&period_start=2026-01-01&period_end=2026-01-31"
        )
        ws = load_workbook(BytesIO(resp.content)).active
        values = _all_values(ws)
        assert any(isinstance(v, str) and "Нет данных" in v for v in values)

    # ── Header block (rows 1–4) ───────────────────────────────────────────────

    def test_project_name_in_row_1(self, client, factories):
        project = factories.ProjectFactory.create(name="Мой Объект")
        resp = client.get(f"/api/export/excel?project_id={project.id}")
        ws = load_workbook(BytesIO(resp.content)).active
        assert ws.cell(row=1, column=1).value == "Мой Объект"

    def test_contract_number_in_row_2(self, client, factories):
        project = factories.ProjectFactory.create(contract_number="Д-2026/01")
        resp = client.get(f"/api/export/excel?project_id={project.id}")
        ws = load_workbook(BytesIO(resp.content)).active
        assert ws.cell(row=2, column=1).value == "Д-2026/01"

    def test_formula_injection_strings_are_escaped(self, client, factories):
        """_safe_str must prefix strings starting with =, +, -, @ with an apostrophe."""
        project = factories.ProjectFactory.create(
            name='=HYPERLINK("http://evil.com")',
            contract_number="+evil",
        )
        mc = factories.MaterialClassFactory.create(name="-В25", calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(
            document=doc, date=date(2026, 3, 10), supplier_name="@inject",
        )
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc, quantity=10.0)
        resp = client.get(
            f"/api/export/excel?project_id={project.id}"
            "&period_start=2026-03-01&period_end=2026-03-31"
        )
        assert resp.status_code == 200
        ws = load_workbook(BytesIO(resp.content)).active
        # Row 1: project name escaped (= → ')
        assert ws.cell(row=1, column=1).value == "'=HYPERLINK(\"http://evil.com\")"
        # Row 2: contract number escaped (+ → ')
        assert ws.cell(row=2, column=1).value == "'+evil"
        # Class section header escaped (- → ')
        assert "'-В25" in _all_values(ws)
        # Data row col C: supplier name escaped (@ → ')
        data_row = _find_data_row(ws)
        assert data_row is not None
        assert ws.cell(row=data_row, column=3).value == "'@inject"

    def test_period_line_in_row_3_when_explicit_period(self, client, factories):
        project = factories.ProjectFactory.create()
        resp = client.get(
            f"/api/export/excel?project_id={project.id}"
            "&period_start=2026-03-01&period_end=2026-03-31"
        )
        ws = load_workbook(BytesIO(resp.content)).active
        assert ws.cell(row=3, column=1).value == "Период: 01.03.2026 — 31.03.2026"

    # ── Column layout ────────────────────────────────────────────────────────

    def test_sheet_has_exactly_21_columns(self, client, factories):
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10))
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc, quantity=10.0)
        resp = client.get(
            f"/api/export/excel?project_id={project.id}"
            "&period_start=2026-03-01&period_end=2026-03-31"
        )
        ws = load_workbook(BytesIO(resp.content)).active
        assert ws.max_column == 21

    def test_formula_columns_contain_excel_formulas(self, client, factories):
        """Columns M(13), N(14), O(15), P(16), Q(17), R(18), S(19) in a data row
        must hold '=...' formula strings (not pre-computed floats)."""
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10))
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc, quantity=10.0)
        resp = client.get(
            f"/api/export/excel?project_id={project.id}"
            "&period_start=2026-03-01&period_end=2026-03-31"
        )
        # data_only=False to read formula strings, not cached values
        ws = load_workbook(BytesIO(resp.content), data_only=False).active
        data_row = _find_data_row(ws)
        assert data_row is not None, "No data row found (col D must have qty > 0)"

        formula_cols = {13: "M", 14: "N", 15: "O", 16: "P", 17: "Q", 18: "R", 19: "S"}
        for col, letter in formula_cols.items():
            v = ws.cell(row=data_row, column=col).value
            assert isinstance(v, str) and v.startswith("="), (
                f"Col {letter}({col}) row {data_row}: expected formula, got {v!r}"
            )

    def test_static_columns_contain_plain_values(self, client, factories):
        """Columns A–L (1–12) in a data row hold plain values, not formulas."""
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10), vat_rate=20.0)
        factories.InvoiceItemFactory.create(
            invoice=inv, material_class=mc,
            quantity=10.0, unit_price=5000.0, amount=50_000.0, vat_amount=10_000.0,
        )
        resp = client.get(
            f"/api/export/excel?project_id={project.id}"
            "&period_start=2026-03-01&period_end=2026-03-31"
        )
        ws = load_workbook(BytesIO(resp.content), data_only=False).active
        data_row = _find_data_row(ws)
        assert data_row is not None

        for col in range(1, 13):  # A–L
            v = ws.cell(row=data_row, column=col).value
            assert not (isinstance(v, str) and v.startswith("=")), (
                f"Col {col} in data row should be plain value, got formula: {v!r}"
            )

    # ── Section structure ────────────────────────────────────────────────────

    def test_class_section_header_appears_in_sheet(self, client, factories):
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(name="В25", calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10))
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc, quantity=10.0)
        resp = client.get(
            f"/api/export/excel?project_id={project.id}"
            "&period_start=2026-03-01&period_end=2026-03-31"
        )
        ws = load_workbook(BytesIO(resp.content)).active
        assert "В25" in _all_values(ws)

    def test_month_header_in_russian_appears_in_sheet(self, client, factories):
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(name="В25", calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10))
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc, quantity=10.0)
        resp = client.get(
            f"/api/export/excel?project_id={project.id}"
            "&period_start=2026-03-01&period_end=2026-03-31"
        )
        ws = load_workbook(BytesIO(resp.content)).active
        assert "Март 2026" in _all_values(ws)

    def test_grand_total_row_per_class_present(self, client, factories):
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(name="В25", calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10))
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc, quantity=10.0)
        resp = client.get(
            f"/api/export/excel?project_id={project.id}"
            "&period_start=2026-03-01&period_end=2026-03-31"
        )
        ws = load_workbook(BytesIO(resp.content)).active
        values = _all_values(ws)
        assert any("ИТОГО по В25" in v for v in values if isinstance(v, str))

    def test_two_classes_produce_two_section_headers_and_grand_totals(self, client, factories):
        project = factories.ProjectFactory.create()
        mc1 = factories.MaterialClassFactory.create(name="В25", calc_role="base")
        mc2 = factories.MaterialClassFactory.create(name="В30", calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10))
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc1, quantity=10.0)
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc2, quantity=5.0)
        resp = client.get(
            f"/api/export/excel?project_id={project.id}"
            "&period_start=2026-03-01&period_end=2026-03-31"
        )
        ws = load_workbook(BytesIO(resp.content)).active
        values = _all_values(ws)
        assert "В25" in values
        assert "В30" in values
        assert any("ИТОГО по В25" in v for v in values if isinstance(v, str))
        assert any("ИТОГО по В30" in v for v in values if isinstance(v, str))

    def test_two_months_produce_thin_separator_row(self, client, factories):
        """When data spans 2 months, there is at least one 6px-height separator row."""
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv_mar = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10))
        inv_apr = factories.InvoiceFactory.create(document=doc, date=date(2026, 4, 5))
        factories.InvoiceItemFactory.create(invoice=inv_mar, material_class=mc, quantity=10.0)
        factories.InvoiceItemFactory.create(invoice=inv_apr, material_class=mc, quantity=10.0)
        resp = client.get(
            f"/api/export/excel?project_id={project.id}"
            "&period_start=2026-03-01&period_end=2026-04-30"
        )
        ws = load_workbook(BytesIO(resp.content)).active
        sep_rows = [r for r in range(1, ws.max_row + 1) if ws.row_dimensions[r].height == 6]
        assert len(sep_rows) >= 1

    def test_single_month_has_no_separator_row(self, client, factories):
        """Single month → no separator rows generated."""
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10))
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc, quantity=10.0)
        resp = client.get(
            f"/api/export/excel?project_id={project.id}"
            "&period_start=2026-03-01&period_end=2026-03-31"
        )
        ws = load_workbook(BytesIO(resp.content)).active
        sep_rows = [r for r in range(1, ws.max_row + 1) if ws.row_dimensions[r].height == 6]
        assert len(sep_rows) == 0

    # ── Filtering ────────────────────────────────────────────────────────────

    def test_period_filter_excludes_out_of_range_invoices(self, client, factories):
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(name="В25", calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv_in = factories.InvoiceFactory.create(
            document=doc, date=date(2026, 3, 10), number="В-периоде"
        )
        inv_out = factories.InvoiceFactory.create(
            document=doc, date=date(2026, 5, 1), number="Вне-периода"
        )
        factories.InvoiceItemFactory.create(invoice=inv_in, material_class=mc, quantity=10.0)
        factories.InvoiceItemFactory.create(invoice=inv_out, material_class=mc, quantity=10.0)
        resp = client.get(
            f"/api/export/excel?project_id={project.id}"
            "&period_start=2026-03-01&period_end=2026-03-31"
        )
        ws = load_workbook(BytesIO(resp.content)).active
        values = _all_values(ws)
        assert "В-периоде" in values
        assert "Вне-периода" not in values

    def test_material_class_id_filter_in_endpoint(self, client, factories):
        project = factories.ProjectFactory.create()
        mc1 = factories.MaterialClassFactory.create(name="В25", calc_role="base")
        mc2 = factories.MaterialClassFactory.create(name="В30", calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10))
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc1, quantity=10.0)
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc2, quantity=5.0)
        resp = client.get(
            f"/api/export/excel?project_id={project.id}"
            f"&period_start=2026-03-01&period_end=2026-03-31"
            f"&material_class_id={mc1.id}"
        )
        ws = load_workbook(BytesIO(resp.content)).active
        values = _all_values(ws)
        assert "В25" in values
        assert "В30" not in values

    def test_no_period_uses_data_bounds(self, client, factories):
        """Without period_start/period_end the endpoint still returns a valid xlsx."""
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(name="В25", calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 4, 15))
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc, quantity=50.0)
        resp = client.get(f"/api/export/excel?project_id={project.id}")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        ws = load_workbook(BytesIO(resp.content)).active
        assert "В25" in _all_values(ws)

    # ── Data values in sheet ─────────────────────────────────────────────────

    def test_excl_vat_value_appears_in_column_j(self, client, factories):
        """mat_per_m3_excl_vat = 5000.0 appears as plain value in col J (10)."""
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10), vat_rate=20.0)
        factories.InvoiceItemFactory.create(
            invoice=inv, material_class=mc,
            quantity=10.0, unit_price=5000.0, amount=50_000.0, vat_amount=10_000.0,
        )
        resp = client.get(
            f"/api/export/excel?project_id={project.id}"
            "&period_start=2026-03-01&period_end=2026-03-31"
        )
        ws = load_workbook(BytesIO(resp.content)).active
        data_row = _find_data_row(ws)
        assert data_row is not None
        assert float(ws.cell(row=data_row, column=10).value) == pytest.approx(5000.0)

    def test_ref_price_appears_in_column_h(self, client, factories):
        project = factories.ProjectFactory.create()
        mc = factories.MaterialClassFactory.create(calc_role="base")
        factories.ReferencePriceFactory.create(
            project=project, material_class=mc, price=7500.0,
            period_start=date(2026, 3, 1), period_end=date(2026, 3, 31),
        )
        doc = factories.DocumentFactory.create(project=project)
        inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10))
        factories.InvoiceItemFactory.create(invoice=inv, material_class=mc, quantity=10.0)
        resp = client.get(
            f"/api/export/excel?project_id={project.id}"
            "&period_start=2026-03-01&period_end=2026-03-31"
        )
        ws = load_workbook(BytesIO(resp.content)).active
        data_row = _find_data_row(ws)
        assert data_row is not None
        assert float(ws.cell(row=data_row, column=8).value) == pytest.approx(7500.0)


# ---------------------------------------------------------------------------
# Section 3: normalized two-block layout (raw + calc columns)
# ---------------------------------------------------------------------------

def test_export_has_raw_and_calc_columns(client, factories, db_session):
    from models import UnitOfMeasure
    project = factories.ProjectFactory.create()
    mc = factories.MaterialClassFactory.create(material_type_code="rebar", name="d12")
    ton = db_session.query(UnitOfMeasure).filter_by(code="TON").one()
    factories.ReferencePriceFactory.create(
        project=project, material_class=mc, unit_id=ton.id, price=60000,
        period_start=date(2026, 1, 1), period_end=date(2026, 12, 31),
    )
    doc = factories.DocumentFactory.create(project=project)
    inv = factories.InvoiceFactory.create(document=doc, date=date(2026, 3, 10))
    # 2000 kg rebar normalized to 2 tons
    factories.InvoiceItemFactory.create(
        invoice=inv, material_class_id=mc.id, raw_unit="кг", quantity=2000,
        normalized_unit_id=ton.id, normalized_quantity=2, unit_price=60, normalized_unit_price=60000,
        amount=120000,
    )
    resp = client.get(
        f"/api/export/excel?project_id={project.id}"
        "&period_start=2026-03-01&period_end=2026-03-31"
    )
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    all_text = {c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)}
    assert "Кол-во по документу" in all_text
    assert "Расчётное кол-во" in all_text
    assert "Базовая ед. изм." in all_text
